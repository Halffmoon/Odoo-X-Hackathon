import io

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.dashboard import (
    BookingHeatmap,
    DepartmentSummaryRow,
    MaintenanceFrequencyRow,
    RetirementForecastRow,
    UtilizationRow,
)


async def utilization(db: AsyncSession) -> list[UtilizationRow]:
    rows = (
        await db.execute(
            text(
                "SELECT a.asset_id, a.asset_tag, a.name AS asset_name, "
                "count(al.allocation_id) AS allocation_count, "
                "COALESCE(SUM(COALESCE(al.actual_return_date, current_date) "
                "- al.allocation_date), 0) AS total_days "
                "FROM assets a "
                "LEFT JOIN asset_allocations al ON al.asset_id = a.asset_id "
                "WHERE a.is_deleted = false "
                "GROUP BY a.asset_id, a.asset_tag, a.name "
                "ORDER BY total_days DESC, allocation_count DESC"
            )
        )
    ).mappings().all()
    return [
        UtilizationRow(
            asset_id=r["asset_id"],
            asset_tag=r["asset_tag"],
            asset_name=r["asset_name"],
            allocation_count=r["allocation_count"],
            total_allocation_days=int(r["total_days"]),
        )
        for r in rows
    ]


async def maintenance_frequency(db: AsyncSession) -> list[MaintenanceFrequencyRow]:
    rows = (
        await db.execute(
            text(
                "SELECT a.asset_id, a.asset_tag, a.category_id, c.name AS category_name, "
                "count(m.maintenance_id) AS request_count, "
                "AVG(EXTRACT(EPOCH FROM (m.resolved_on - m.created_on)) / 3600.0) "
                "FILTER (WHERE m.resolved_on IS NOT NULL) AS avg_hours "
                "FROM assets a "
                "JOIN maintenance_requests m ON m.asset_id = a.asset_id "
                "LEFT JOIN asset_categories c ON c.category_id = a.category_id "
                "WHERE a.is_deleted = false "
                "GROUP BY a.asset_id, a.asset_tag, a.category_id, c.name "
                "ORDER BY request_count DESC"
            )
        )
    ).mappings().all()
    return [
        MaintenanceFrequencyRow(
            asset_id=r["asset_id"],
            asset_tag=r["asset_tag"],
            category_id=r["category_id"],
            category_name=r["category_name"],
            request_count=r["request_count"],
            avg_resolution_hours=(
                round(float(r["avg_hours"]), 2) if r["avg_hours"] is not None else None
            ),
        )
        for r in rows
    ]


async def department_summary(db: AsyncSession) -> list[DepartmentSummaryRow]:
    rows = (
        await db.execute(
            text(
                "SELECT d.department_id, d.name AS department_name, "
                "count(a.asset_id) AS total_assets, "
                "count(*) FILTER (WHERE a.status='ALLOCATED') AS allocated, "
                "count(*) FILTER (WHERE a.status='AVAILABLE') AS available, "
                "count(*) FILTER (WHERE a.status='UNDER_MAINTENANCE') AS under_maintenance "
                "FROM departments d "
                "LEFT JOIN assets a ON a.current_department_id = d.department_id "
                "AND a.is_deleted = false "
                "WHERE d.is_deleted = false "
                "GROUP BY d.department_id, d.name "
                "ORDER BY d.name"
            )
        )
    ).mappings().all()
    return [
        DepartmentSummaryRow(
            department_id=r["department_id"],
            department_name=r["department_name"],
            total_assets=r["total_assets"],
            allocated=r["allocated"],
            available=r["available"],
            under_maintenance=r["under_maintenance"],
        )
        for r in rows
    ]


async def booking_heatmap(db: AsyncSession) -> BookingHeatmap:
    rows = (
        await db.execute(
            text(
                "SELECT EXTRACT(DOW FROM start_time)::int AS dow, "
                "EXTRACT(HOUR FROM start_time)::int AS hr, count(*) AS cnt "
                "FROM bookings WHERE status <> 'CANCELLED' "
                "GROUP BY dow, hr"
            )
        )
    ).mappings().all()
    matrix = [[0 for _ in range(24)] for _ in range(7)]
    for r in rows:
        matrix[r["dow"]][r["hr"]] = r["cnt"]
    return BookingHeatmap(matrix=matrix)


async def retirement_forecast(
    db: AsyncSession, older_than_years: int = 5
) -> list[RetirementForecastRow]:
    rows = (
        await db.execute(
            text(
                "SELECT asset_id, asset_tag, name AS asset_name, condition, acquisition_date "
                "FROM assets WHERE is_deleted = false AND ("
                "condition IN ('POOR','DAMAGED') OR "
                "(acquisition_date IS NOT NULL AND acquisition_date < "
                "current_date - make_interval(years => :yrs))) "
                "ORDER BY acquisition_date NULLS LAST"
            ),
            {"yrs": older_than_years},
        )
    ).mappings().all()
    out: list[RetirementForecastRow] = []
    for r in rows:
        reasons = []
        if r["condition"] in ("POOR", "DAMAGED"):
            reasons.append(f"condition is {r['condition']}")
        if r["acquisition_date"] is not None:
            reasons.append(f"acquired {r['acquisition_date']}")
        out.append(
            RetirementForecastRow(
                asset_id=r["asset_id"],
                asset_tag=r["asset_tag"],
                asset_name=r["asset_name"],
                condition=r["condition"],
                acquisition_date=r["acquisition_date"],
                reason="; ".join(reasons),
            )
        )
    return out


# ---------------- Export ----------------

_EXPORTERS = {
    "utilization": (
        utilization,
        ["asset_tag", "asset_name", "allocation_count", "total_allocation_days"],
    ),
    "maintenance-frequency": (
        maintenance_frequency,
        ["asset_tag", "category_name", "request_count", "avg_resolution_hours"],
    ),
    "department-summary": (
        department_summary,
        ["department_name", "total_assets", "allocated", "available", "under_maintenance"],
    ),
    "retirement-forecast": (
        retirement_forecast,
        ["asset_tag", "asset_name", "condition", "acquisition_date", "reason"],
    ),
}


async def export_report(
    db: AsyncSession, report_type: str, fmt: str
) -> tuple[bytes, str, str]:
    """Return (bytes, media_type, filename)."""
    if report_type not in _EXPORTERS:
        raise ValueError(f"Unknown report type '{report_type}'.")
    fn, columns = _EXPORTERS[report_type]
    rows = await fn(db)
    dict_rows = [r.model_dump() for r in rows]

    if fmt == "xlsx":
        return _to_xlsx(report_type, columns, dict_rows)
    if fmt == "pdf":
        return _to_pdf(report_type, columns, dict_rows)
    raise ValueError("format must be 'xlsx' or 'pdf'.")


def _to_xlsx(report_type, columns, rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = report_type[:31]
    ws.append([c.replace("_", " ").title() for c in columns])
    for row in rows:
        ws.append([_cell(row.get(c)) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{report_type}.xlsx",
    )


def _to_pdf(report_type, columns, rows):
    from fpdf import FPDF

    pdf = FPDF(orientation="L")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, report_type.replace("-", " ").title(), ln=1)
    pdf.set_font("Helvetica", "B", 9)
    width = 277 / len(columns)
    for c in columns:
        pdf.cell(width, 8, c.replace("_", " ").title(), border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    for row in rows:
        for c in columns:
            pdf.cell(width, 7, str(_cell(row.get(c)))[:40], border=1)
        pdf.ln()
    data = pdf.output(dest="S")
    if isinstance(data, str):
        data = data.encode("latin-1")
    return bytes(data), "application/pdf", f"{report_type}.pdf"


def _cell(value):
    if value is None:
        return ""
    return value
